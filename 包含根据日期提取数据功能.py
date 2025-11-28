import gradio as gr
import threading
import requests
import json
import lark_oapi as lark
from lark_oapi.api.bitable.v1 import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
import traceback
from datetime import datetime, timedelta
from openai import OpenAI
import time
import sys
import io
from urllib.parse import urlparse, unquote

# --- 修复Windows控制台编码问题 ---
if sys.stdout and sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ==============================================================================
# --- 1. 全局配置 ---
# ==============================================================================

# --- 飞书应用凭证 ---
LARK_APP_ID = os.getenv('LARK_APP_ID')
LARK_APP_SECRET = os.getenv('LARK_APP_SECRET')
LARK_APP_TOKEN = os.getenv('LARK_APP_TOKEN')

# --- 飞书多维表格ID ---
# 月度
LARK_TABLE_ID_MONTH_CURRENT = 'tblcD6vnYng9Cqim' # 10月
LARK_TABLE_ID_MONTH_PREVIOUS = 'tbl0GqfuwXAvIwsT' # 9月
LARK_TABLE_ID_PRODUCT_MONTH_CURRENT = 'tblHY5xRjblqoPGz' # 10月
LARK_TABLE_ID_PRODUCT_MONTH_PREVIOUS = 'tblkOmN2EiOFeVHY' # 9月
# 周度
LARK_TABLE_ID_WEEK_CURRENT = 'tbluVbrXLRUmfouv'
LARK_TABLE_ID_WEEK_PREVIOUS = 'tblEcILnU5J0N6JC'
LARK_TABLE_ID_PRODUCT_WEEK_CURRENT = 'tblAJIhA78vK5Tpj'
LARK_TABLE_ID_PRODUCT_WEEK_PREVIOUS = 'tblPV2YRW7JFt1F0'


# --- AI模型配置 ---
MODELSCOPE_API_KEY = os.getenv('MODELSCOPE_API_KEY')
MODELSCOPE_BASE_URL = 'https://gemini.zzh2025.dpdns.org'
DATA_ANALYSIS_MODEL_ID = 'gemini-2.5-flash-lite'

# --- 全局日志记录 ---
log_lock = threading.Lock()
GLOBAL_LOGS = []

def log_print(*args, **kwargs):
    """带线程锁的日志打印函数，将日志记录到全局列表并打印到控制台。"""
    with log_lock:
        message = " ".join(str(arg) for arg in args)
        print(message, **kwargs, flush=True) # 仍然打印到控制台，方便调试
        GLOBAL_LOGS.append(message)

# ==============================================================================
# --- 2. 数据下载和处理功能 ---
# ==============================================================================

def download_weekly_data(start_date_str, end_date_str, data_source="life_data"):
    """
    从生活服务API下载指定日期范围的周数据Excel文件
    """
    log_print(f"--- LOG: 开始下载周数据 ({start_date_str} to {end_date_str})...")
    
    api_url = 'https://www.life-data.cn/api/dito/query'
    
    passport_csrf_token = os.getenv('PASSPORT_CSRF_TOKEN', '86dc5731da9047f4e92be57f55951317')
    passport_csrf_token_default = os.getenv('PASSPORT_CSRF_TOKEN_DEFAULT', '86dc5731da9047f4e92be57f55951317')
    
    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6',
        'cache-control': 'no-cache',
        'content-type': 'application/json',
        'cookie': f'passport_csrf_token={passport_csrf_token}; passport_csrf_token_default={passport_csrf_token_default}; is_staff_user=false; sid_guard=cdeed2a82c0c946ac3639cb94a137341%7C1761719939%7C5012576%7CFri%2C+26-Dec-2025+07%3A01%3A55+GMT; uid_tt=b28e76efc6fbcfc77440b989e9511d65; uid_tt_ss=b28e76efc6fbcfc77440b989e9511d65; sid_tt=cdeed2a82c0c946ac3639cb94a137341; sessionid=cdeed2a82c0c946ac3639cb94a137341; sessionid_ss=cdeed2a82c0c946ac3639cb94a137341; session_tlb_tag=sttt%7C5%7Cze7SqCwMlGrDY5y5ShNzQf________-sz6tFDOQ2k8HIpCnXO1msPRYK0G2iEZN8xOK9ASBGKiM%3D; sid_ucp_v1=1.0.0-KGU3YzQ5NzgzOGYwNDk5YTA1MDlkMjMyNzc4YWZmOWQ1Yzk5OTMxOWYKGAj-xLC9_cykAhCD7YbIBhjMrB04AUDrBxoCaGwiIGNkZWVkMmE4MmMwYzk0NmFjMzYzOWNiOTRhMTM3MzQx; ssid_ucp_v1=1.0.0-KGU3YzQ5NzgzOGYwNDk5YTA1MDlkMjMyNzc4YWZmOWQ1Yzk5OTMxOWYKGAj-xLC9_cykAhCD7YbIBhjMrB04AUDrBxoCaGwiIGNkZWVkMmE4MmMwYzk0NmFjMzYzOWNiOTRhMTM3MzQx; csrf_session_id=8c5e5b4ef482b024d75764477243fe85; gd_random=eyJtYXRjaCI6dHJ1ZSwicGVyY2VudCI6MC43MzkzNDMyMTU3MTYwNjg0fQ==.eOduivmuTmhYGoUg31jlCi02FuQ4WtFtJW71nnCYlJk=',
        'life-account-id': '7241078611527075855',
        'origin': 'https://www.life-data.cn',
        'pragma': 'no-cache',
        'priority': 'u=1, i',
        'referer': 'https://www.life-data.cn/store/my/chain/poi/overview?groupid=1768205901316096',
        'related-account-id': '0',
        'root-life-account-id': '7241078611527075855',
        'sec-ch-ua': '"Microsoft Edge";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36 Edg/141.0.0.0',
        'x-secsdk-csrf-token': '00010000000196abe940802a1054792d8960210611edc788c509574afb23bc5141c9d52b08a01877834669b4b3be',
        'x-tt-ls-session-id': 'a073b46d-a358-4c97-b8aa-359fd5b2a720',
        'x-tt-trace-id': '00-7c4f78ce18dfba45305d491cb-7c4f78ce18dfba45-01',
        'x-tt-trace-log': '01',
    }
    
    payload = {
        "biz_params": {
            "path": "/store/my/chain/poi/overview",
            "first_render": False,
            "common_params": {
                "end_date": end_date_str,
                "date_type": "custom",
                "start_date": start_date_str
            },
            "module_params": {
                "AllPoiList": {
                    "poi_id": [],
                    "brand_id": [],
                    "poi_type": [],
                    "poi_sizer": {},
                    "indicators": [
                        "enter_poi_uv_cnt", "visit_deal_uv_convert_rate", "pay_intention_gmv_1d",
                        "verify_amount_1d", "verify_user_cnt_1d", "verify_cert_cnt_1d",
                        "verify_new_user_cnt_1d", "verify_old_user_cnt_1d", "poi_score",
                        "manage_score", "positive_rate_cnt_1d", "normal_negative_rate",
                        "consumption_rate_cnt_1d", "enter_poi_avg_cnt", "click_poi_project_card_cnt_1d",
                        "click_poi_project_card_uv_cnt_1d", "pay_user_cnt_1d", "pay_cert_cnt_1d",
                        "visit_deal_convert_rate", "enter_poi_cnt", "video_cnt_1d",
                        "video_play_cnt_1d", "convert_label", "pay_gmv", "refund_amount",
                        "refund_cert_cnt", "refund_user_cnt", "new_rate_cnt_1d", "reply_rate_ratio",
                        "bad_comment_ratio", "cs_ticket_ratio", "account_refund_order_ratio",
                        "visible_checkin_cnt_1d", "visible_checkin_item_cnt_1d", "favorite_cnt_1d",
                        "pay_intention_cert_cnt_1d", "pay_intention_user_cnt_1d",
                        "refund_intention_gmv", "refund_intention_cert_cnt",
                        "refund_intention_user_cnt", "rank_text"
                    ],
                    "download": 1
                }
            }
        }
    }
    
    try:
        # --- Step 1: Initiate download task and get task_id ---
        log_print("步骤1/2: 正在请求创建导出任务...")
        initial_response = requests.post(api_url, headers=headers, json=payload, timeout=30)
        initial_response.raise_for_status()
        initial_data = initial_response.json()

        # Extract task_id from the first response
        task_id = initial_data.get('data', [{}])[0].get('task_id')
        if not task_id:
            log_print(f"解析失败！未能从初始响应中找到 'task_id'。响应: {initial_data}")
            return None
        log_print(f"成功创建导出任务，任务ID: {task_id}")

        # --- Step 2: Use task_id to poll for the download URL ---
        log_print("步骤2/2: 等待5秒后，使用任务ID获取下载链接...")
        time.sleep(5)

        # Create payload for the second request, including the task_id
        second_payload = json.loads(json.dumps(payload)) # Deep copy
        second_payload['biz_params']['module_params']['AllPoiList']['task_id'] = task_id
        
        final_response = requests.post(api_url, headers=headers, json=second_payload, timeout=30)
        final_response.raise_for_status()
        final_data = final_response.json()
        
        # Extract download URL from the second response
        download_url = final_data.get('data', [{}])[0].get('url')
        if not download_url:
            log_print(f"解析失败！未能从最终响应中找到下载链接 'url'。响应: {final_data}")
            return None
        log_print(f"成功提取下载链接: {download_url}")

        # --- Step 3: Download the file ---
        log_print("正在下载Excel文件...")
        download_response = requests.get(download_url, timeout=60) # Increased timeout for download
        download_response.raise_for_status()
        
        # Generate filename and save
        file_name = f"周数据_{start_date_str}_{end_date_str}.xlsx"
        
        with open(file_name, 'wb') as f:
            f.write(download_response.content)
        
        log_print(f"下载完成！文件已保存为: {file_name}")
        return file_name
        
    except requests.exceptions.RequestException as e:
        log_print(f"API请求失败: {e}")
        return None
    except (KeyError, IndexError, TypeError) as e:
        log_print(f"解析API响应失败: {e}. 请检查API返回的数据结构。")
        traceback.print_exc()
        return None
    except Exception as e:
        log_print(f"下载过程中发生未知错误: {e}")
        traceback.print_exc()
        return None

def parse_excel_data(file_path, store_name):
    """
    解析下载的Excel文件，提取指定门店的数据
    """
    log_print(f"--- LOG: 开始解析Excel文件 {file_path}，查找门店: {store_name}")
    
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path)
        log_print(f"--- LOG: Excel文件包含 {len(df)} 行数据")
        
        # 查找指定门店的数据
        # 优先查找精确匹配"门店名称"的列
        store_column = None
        for col in df.columns:
            if str(col).strip() == '门店名称':
                store_column = col
                break
        
        # 如果没有找到精确匹配的"门店名称"列，则查找包含"门店"的列
        if not store_column:
            for col in df.columns:
                if '门店' in str(col) or '店铺' in str(col):
                    store_column = col
                    break
        
        if not store_column:
            log_print("--- ERROR: 未找到门店名称列")
            return None
        
        log_print(f"--- LOG: 找到门店名称列: '{store_column}'")
        
        # 筛选指定门店的数据
        # 使用精确匹配而不是contains
        store_data = df[df[store_column].astype(str).str.strip() == store_name.strip()]
        
        if store_data.empty:
            # 如果精确匹配失败，尝试使用contains
            store_data = df[df[store_column].astype(str).str.contains(store_name, na=False)]
            if store_data.empty:
                log_print(f"--- WARNING: 未找到门店 '{store_name}' 的数据")
                return None
        
        log_print(f"--- LOG: 找到门店 '{store_name}' 的数据，共 {len(store_data)} 行")
        
        # 将数据转换为字典格式，保持与原飞书API数据结构一致
        result_data = {}
        for col in store_data.columns:
            if len(store_data) > 0:
                result_data[col] = store_data[col].iloc[0]
        
        return result_data
        
    except Exception as e:
        log_print(f"--- ERROR: 解析Excel文件时发生错误: {e}")
        traceback.print_exc()
        return None

# ==============================================================================
# --- 3. 核心数据处理逻辑 (与原文件相同) ---
# ==============================================================================

def get_store_info_from_feishu(store_name, client) -> dict:
    """
    新功能：根据飞书官方文档，使用 search 接口在线查询门店ID。
    """
    log_print(f"--- LOG: 正在从飞书在线查询 '{store_name}' 的门店ID")
    # 注意：为了通用性，这里仍然使用月度表来查询门店ID，因为周度表中可能不包含所有门店
    request_body = SearchAppTableRecordRequestBody.builder() \
        .filter({
            "conjunction": "and",
            "conditions": [
                {
                    "field_name": "门店名称",
                    "operator": "is",
                    "value": [store_name]
                }
            ]
        }) \
        .build()
    request = SearchAppTableRecordRequest.builder() \
        .app_token(LARK_APP_TOKEN) \
        .table_id(LARK_TABLE_ID_MONTH_CURRENT) \
        .page_size(1) \
        .request_body(request_body) \
        .build()
    try:
        response = client.bitable.v1.app_table_record.search(request)
        if response.success() and response.data.items:
            store_id = response.data.items[0].fields.get("门店ID")
            if store_id:
                return {"门店ID": store_id}
        return None
    except Exception as e:
        log_print(f"--- EXCEPTION during store ID query: {e}")
        traceback.print_exc()
        return None

def query_data_from_table(client: lark.Client, table_id: str, filter_query: str, name: str) -> list:
    """通用的飞书多维表格数据查询函数。"""
    all_records = []
    page_token = None
    retries = 3
    log_print(f"--- LOG: 开始从 Table ID: {table_id} ({name}) 查询数据，筛选条件: {filter_query}")
    while True:
        try:
            request_builder = ListAppTableRecordRequest.builder() \
                .app_token(LARK_APP_TOKEN) \
                .table_id(table_id) \
                .filter(filter_query) \
                .page_size(500)
            if page_token:
                request_builder.page_token(page_token)
            
            request = request_builder.build()
            response = client.bitable.v1.app_table_record.list(request)

            if not response.success():
                log_print(f"--- ERROR: 查询飞书表格 ({name}) 失败: {response.code} {response.msg}")
                if response.code == 1254607 and retries > 0:
                    retries -= 1
                    log_print(f"--- WARNING: 数据尚未准备好，将在3秒后重试... (剩余次数: {retries})")
                    time.sleep(3)
                    continue
                break
            
            items = response.data.items or []
            for item in items:
                all_records.append(item.fields)
            
            if response.data.has_more:
                page_token = response.data.page_token
            else:
                break
        except Exception as e:
            log_print(f"--- EXCEPTION during data query from {table_id} ({name}): {e}")
            traceback.print_exc()
            break
    log_print(f"--- LOG: 从 Table ID: {table_id} ({name}) 查询到 {len(all_records)} 条记录。")
    return all_records

def generate_product_insights_with_ai(raw_sales_records: list) -> list:
    """使用AI模型聚合商品数据。"""
    log_print("\n--- LOG: [AI商品分析] 开始调用LLM处理原始销售数据...")
    if not raw_sales_records:
        return []
    client = OpenAI(base_url=MODELSCOPE_BASE_URL, api_key=MODELSCOPE_API_KEY)
    prompt = f"""
    你是一位顶级数据分析师。请处理以下原始JSON数据，它来自飞书表格。
    **原始数据:**
    ```json
    {json.dumps(raw_sales_records, ensure_ascii=False, indent=2)}
    ```
    ---
    **处理任务:**
    1.  **解析字段**: 
        - `商品名称` 字段可能是 `[{{'text': '名称'}}]` 或纯文本，请提取出文本。
        - `转化率` 字段是 `[{{'text': 'xx.xx%'}}]`，请提取出百分比文本。
        - `实付总金额` 和 `核销次数` 是数值。
    2.  **分组聚合**: 按解析出的`商品名称`进行分组。
        - `amount`: 累加每个商品所有记录的`实付总金额`。
        - `count`: 累加每个商品所有记录的`核销次数`。
        - `conversion_rate`: 对于每个商品，只保留其所有记录中**第一个出现**的`转化率`文本值。
    3.  **输出JSON**: 返回一个聚合后的JSON数组，每个对象包含`name`, `amount`, `count`, `conversion_rate`四个字段。
    **输出要求:**
    - 只输出纯粹的、不含任何其他文字说明的JSON数组。
    """
    try:
        response = client.chat.completions.create(model=DATA_ANALYSIS_MODEL_ID, messages=[{"role": "system", "content": "你是一位精通JSON处理和数据聚合的专家。"},{"role": "user", "content": prompt}], temperature=0.0)
        result_text = response.choices[0].message.content
        cleaned_json = re.sub(r'```json\s*|\s*```', '', result_text.strip())
        final_data = json.loads(cleaned_json)
        log_print(f"--- LOG: [AI商品分析] 成功处理了 {len(final_data)} 条商品数据。")
        return final_data
    except Exception as e:
        log_print(f"--- EXCEPTION: [AI商品分析] 调用AI或解析其响应时发生错误: {e}")
        return []

def get_loss_data_from_api(poi_id, start_date_str, end_date_str):
    """
    根据POI ID和日期范围，从生活服务API获取流失数据。
    """
    log_print(f"--- LOG: 开始从API获取POI ID: {poi_id} 的流失数据 ({start_date_str} to {end_date_str})...")
    url = 'https://www.life-data.cn/api/dito/query'

    passport_csrf_token = os.getenv('PASSPORT_CSRF_TOKEN')
    passport_csrf_token_default = os.getenv('PASSPORT_CSRF_TOKEN_DEFAULT')

    if not passport_csrf_token or not passport_csrf_token_default:
        log_print("--- ERROR: 缺少环境变量 PASSPORT_CSRF_TOKEN 或 PASSPORT_CSRF_TOKEN_DEFAULT。无法获取流失数据。")
        return None

    # 注意：x-secsdk-csrf-token 和 cookie 中的其他部分是硬编码的，可能会过期
    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6',
        'content-type': 'application/json',
        'cookie': f'passport_csrf_token={passport_csrf_token}; passport_csrf_token_default={passport_csrf_token_default}; is_staff_user=false; sid_guard=9ea139c5cf82152a939f27fbffcea265%7C1757900485%7C4937074%7CTue%2C+11-Nov-2025+05%3A05%3A59+GMT; uid_tt=385abdb04a651858a9498813db9b1997; uid_tt_ss=385abdb04a651858a9498813db9b1997; sid_tt=9ea139c5cf82152a939f27fbffcea265; sessionid=9ea139c5cf82152a939f27fbffcea265; sessionid_ss=9ea139c5cf82152a939f27fbffcea265; session_tlb_tag=sttt%7C7%7CnqE5xc-CFSqTnyf7_86iZf________-qu5u4YHHUo1u9c67rioxFeEGaOndSMj2sn8qaulgNbkQ%3D; sid_ucp_v1=1.0.0-KDE3NjU5NmE3MTY1MjdiZGE3YjI1MDFiOTAwNGUwOGZmMWYwZmI0MzgKGAj-xLC9_cykAhDF3Z3GBhjMrB04AUDrBxoCbGYiIDllYTEzOWM1Y2Y4MjE1MmE5MzlmMjdmYmZmY2VhMjY1; ssid_ucp_v1=1.0.0-KDE3NjU5NmE3MTY1MjdiZGE3YjI1MDFiOTAwNGUwOGZmMWYwZmI0MzgKGAj-xLC9_cykAhDF3Z3GBhjMrB04AUDrBxoCbGYiIDllYTEzOWM1Y2Y4MjE1MmE5MzlmMjdmYmZmY2VhMjY1; csrf_session_id=2bcc7607e102e9e083c7c77643165d0d; gd_random=eyJtYXRjaCI6dHJ1ZSwicGVyY2VudCI6MC40NjAyMjc0MzczNTU1NDAzfQ==.deVjVQ0KiEmH8xLXM+lRVKMPdkMz8Yy/D5GaRCxRrM4=',
        'life-account-id': '7241078611527075855',
        'origin': 'https://www.life-data.cn',
        'referer': f'https://www.life-data.cn/store/my/chain/list/details?poi_id={poi_id}',
        'root-life-account-id': '7241078611527075855',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0',
        'x-secsdk-csrf-token': '0001000000012e648d9e7809dd66cc87e29291e6f35ade75bdac56ba417906efc7f3ed50fdb518655820c59e13c9',
    }

    payload = {
        "biz_params": {
            "path": "/poi/detail/v2", "query": {}, "first_render": False,
            "common_params": {
                "start_date": start_date_str, "end_date": end_date_str,
                "date_type": "custom", "poi_id": str(poi_id)
            },
            "module_params": {"LossAnalysis": {}}
        },
        "dito_params": {
            "is_event": True,
            "node_update_map": [{"type": "refresh", "node": "PcStoreLossAnalysis1"}]
        }
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=20)
        response.raise_for_status()
        log_print(f"--- SUCCESS: API请求成功 (POI ID: {poi_id}, 状态码: {response.status_code})")
        response_data = response.json()
        if response_data.get('code') == 0:
            layout = response_data.get('data', {}).get('layout', [])
            for component in layout:
                if component.get('id') == 'PcStoreLossAnalysis1':
                    loss_data_list = component.get('data', {}).get('LossAnalysis', {}).get('data', [])
                    if loss_data_list:
                        loss_data = loss_data_list[0]
                        log_print(f"--- LOG: 成功从API解析到流失数据: {loss_data}")
                        return {
                            "api_visit_lost_count": loss_data.get('valid_enter_poi_lost_uv_cnt'),
                            "api_lost_to_other_store_deal_count": loss_data.get('lost_pay_user_cnt')
                        }
            log_print(f"--- WARNING: API返回数据中未找到 'PcStoreLossAnalysis1' 组件。")
            return None
        log_print(f"--- WARNING: API返回数据格式不正确或code不为0: {response_data.get('code')} {response_data.get('message')}")
        return None
    except requests.exceptions.HTTPError as errh:
        log_print(f"--- EXCEPTION: API请求发生HTTP错误: {errh}. Response: {response.text}")
    except requests.exceptions.RequestException as e:
        log_print(f"--- EXCEPTION: API请求失败: {e}")
    return None

def get_field_value(record, field_name, default=None):
    """安全地从飞书记录中提取字段值。"""
    if not record or field_name not in record:
        return default
    value = record[field_name]
    if isinstance(value, list) and value:
        if isinstance(value[0], dict) and 'text' in value[0]:
            return value[0]['text']
    if isinstance(value, str):
        try:
            cleaned_value = value.replace(',', '')
            if '.' in cleaned_value:
                return float(cleaned_value)
            return int(cleaned_value)
        except (ValueError, TypeError):
            return value
    return value if value is not None else default

def calculate_percentage_change(current, previous):
    """计算环比变化。"""
    if previous is None or current is None or str(previous).strip() == '' or str(current).strip() == '':
        return None
    try:
        current = float(str(current).replace('%', ''))
        previous = float(str(previous).replace('%', ''))
    except (ValueError, TypeError):
        return "N/A"
    if previous == 0:
        return "∞" if current > 0 else "N/A"
    change = ((current - previous) / abs(previous)) * 100
    arrow = "↑ " if change > 0 else "↓ "
    return f"{arrow}{abs(change):.2f}%"

def generate_ai_analysis(data_summary, analysis_type):
    """使用AI模型生成数据分析报告。"""
    log_print(f"\n--- LOG: [AI{analysis_type}] 开始调用LLM进行数据分析...")
    client = OpenAI(base_url=MODELSCOPE_BASE_URL, api_key=MODELSCOPE_API_KEY)
    if analysis_type == "problem_analysis":
        prompt = f"""
        你是一位资深的数据分析师和商业顾问。请根据以下门店经营数据，进行深入的问题分析。
        **门店数据摘要：**
        ```json
        {json.dumps(data_summary, ensure_ascii=False, indent=2, default=str)}
        ```
        **分析要求：**
        1. 重点关注数据中的异常值、负增长趋势和潜在问题
        2. 结合网吧行业特点，分析数据背后的业务含义
        3. 数据来源如果是英文字段，要进行翻译成中文的操作，相关指标不能是英文的
        **输出要求：**
        - 直接输出问题分析内容，不要包含标题、前言或其他格式
        - 内容要简洁明了，重点突出
        - 严格控制在50字以内，一段话阐述清楚主要问题
        - 不要使用星号（*）等特殊符号来进行内容标注或强调
        - 不要引用具体的数据百分比，只描述问题本质
        """
    else:  # improvement_suggestions
        prompt = f"""
        你是一位资深的数据分析师和商业顾问。请根据以下门店经营数据，提供具体的改进建议。
        **门店数据摘要：**
        ```json
        {json.dumps(data_summary, ensure_ascii=False, indent=2, default=str)}
        ```
        **建议要求：**
        1. 针对数据分析中发现的问题，提供具体可行的改进建议
        2. 结合电竞行业特点，提供通用性的建议，避免提及具体套餐名称
        **输出要求：**
        - 直接输出改进建议内容，不要包含标题、前言或其他格式
        - 内容要实用性强，重点突出
        - 严格控制在50字以内
        - 不要使用星号（*）等特殊符号来进行内容标注或强调
        - 不要提及技术术语如"api_visit_lost_count"，使用业务语言
        - 不要提及具体的套餐名称
        """
    try:
        response = client.chat.completions.create(
            model=DATA_ANALYSIS_MODEL_ID,
            messages=[
                {"role": "system", "content": "你是一位精通数据分析和商业咨询的专家，擅长从数据中发现问题并提供解决方案。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        result_text = response.choices[0].message.content.strip()
        log_print(f"--- LOG: [AI{analysis_type}] 成功生成分析内容")
        return result_text
    except Exception as e:
        log_print(f"--- EXCEPTION: [AI{analysis_type}] 调用AI或解析其响应时发生错误: {e}")
        return "分析生成失败，请稍后重试。"

def write_product_analysis_sheet(wb, all_data, store_name, sort_by="销售额", analysis_type="月度分析"):
    """将套餐数据写入“套餐分析”sheet页，支持月度和周度，按指定字段排序，并根据转化率进行着色，同时设置居中和边框。"""
    try:
        log_print(f"\n--- LOG: [套餐分析] 开始填充'套餐分析'Sheet页 (排序方式: {sort_by}, 分析类型: {analysis_type})...")
        ws = wb["套餐分析"]
        log_print("--- DEBUG: [套餐分析] 已获取'套餐分析'工作表。")

        # --- 设置大标题 ---
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{store_name} 套餐数据分析报告"
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(bold=True, size=20)
        title_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        log_print("--- DEBUG: [套餐分析] 标题已成功写入并格式化。")

        # --- 动态设置子表头 ---
        current_period_header = "本周套餐" if analysis_type == "周分析" else "10月套餐"
        previous_period_header = "上周套餐" if analysis_type == "周分析" else "9月套餐"
        ws.merge_cells('A2:D2')
        ws['A2'].value = current_period_header
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('E2:H2')
        ws['E2'].value = previous_period_header
        ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E2'].font = Font(bold=True)
        log_print(f"--- DEBUG: [套餐分析] 子表头已更新为: '{current_period_header}' 和 '{previous_period_header}'")

        alignment_center = Alignment(horizontal='center', vertical='center')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        color_low_sales = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        color_medium_sales = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
        color_high_sales = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

        def process_and_sort_product_data(raw_data, sort_key):
            processed_list = [
                {
                    "套餐名称": get_field_value(record, "商品名称", "N/A"),
                    "销售额": get_field_value(record, "实付总金额", 0),
                    "核销次数": get_field_value(record, "核销次数", 0),
                    "转化率": get_field_value(record, "转化率", "0%")
                } for record in raw_data
            ]
            df = pd.DataFrame(processed_list)
            if not df.empty and sort_key in df.columns:
                df = df.sort_values(by=sort_key, ascending=False).reset_index(drop=True)
            return df

        df_current = process_and_sort_product_data(all_data.get('product_sales_data_current', []), sort_by)
        df_previous = process_and_sort_product_data(all_data.get('product_sales_data_previous', []), sort_by)

        if ws.max_row > 3: # 从第4行开始清除
            ws.delete_rows(4, ws.max_row)

        def write_and_format_data(df, start_col):
            # 从第4行开始写数据
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
                fill_color = None
                try:
                    # 根据“核销次数”进行颜色编码
                    sales_volume = int(row_data[2])
                    if 1 <= sales_volume <= 10:
                        fill_color = color_low_sales
                    elif 11 <= sales_volume <= 20:
                        fill_color = color_medium_sales
                    elif sales_volume > 20:
                        fill_color = color_high_sales
                except (ValueError, TypeError, IndexError):
                    pass
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx + start_col, value=value)
                    cell.alignment = alignment_center
                    cell.border = border_thin
                    if fill_color:
                        cell.fill = fill_color
        
        write_and_format_data(df_current, start_col=0)
        write_and_format_data(df_previous, start_col=4)

        for col_cells in ws.columns:
            max_length = 0
            column = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    length = len(str(cell.value).encode('gbk')) if re.search("[\u4e00-\u9fa5]", str(cell.value)) else len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except: pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width if adjusted_width < 60 else 60

        log_print("--- SUCCESS: [套餐分析] '套餐分析'Sheet页填充和格式化完毕。")

    except KeyError:
        log_print("--- WARNING: [套餐分析] 模板中未找到名为 '套餐分析' 的Sheet页，已跳过此步骤。")
    except Exception as e:
        log_print(f"--- EXCEPTION: [套餐分析] 填充Sheet页时发生错误: {e}")
        traceback.print_exc()

def generate_report(store_name, store_id, all_data, sort_by, analysis_type):
    """根据模板生成最终的Excel报告 (重构版，不使用pandas读取)。"""
    log_print(f"\n--- LOG: [报告生成] 准备进入generate_report函数 (分析类型: {analysis_type})...")
    template_path = 'AI数据分析模版.xlsx'
    safe_store_name = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '_', store_name)
    output_filename = f"{safe_store_name}_分析报告_{datetime.now().strftime('%Y%m%d')}.xlsx"
    try:
        import shutil
        shutil.copy2(template_path, output_filename)
        wb = load_workbook(output_filename)
        
        # --- 1. 处理“数据分析”工作表 ---
        ws = wb["数据分析"]
        log_print("--- DEBUG: [数据分析] 已获取'数据分析'工作表。")
        
        # 设置大标题
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{store_name} 经营数据分析报告"
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(bold=True, size=20)
        title_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        log_print("--- DEBUG: [数据分析] 标题已成功写入并格式化。")
        
        # --- 根据分析类型动态修改表头 ---
        current_period_col_name = "本周数值" if analysis_type == "周分析" else "10月数值"
        previous_period_col_name = "上周数值" if analysis_type == "周分析" else "9月数值"
        ws['C2'].value = current_period_col_name
        ws['D2'].value = previous_period_col_name
        log_print(f"--- DEBUG: [数据分析] 表头已更新为: {current_period_col_name}, {previous_period_col_name}")


        data_current = all_data.get('current_period_data', {})
        data_previous = all_data.get('previous_period_data', {})
        
        field_mapping = {
            "门店曝光量": "门店页访问次数", "门店详情页浏览量": "门店页访问人数", "门店页成交人数": "门店意向成交人数",
            "新增评论数": "新增评价数", "评价回复率": "评价回复率", "商责退单率": "经营风险商责退单率",
            "团购券购买数": "门店意向成交券数", "团购券购买金额": "门店意向成交金额", "团购券核销数": "门店核销券数",
            "核销金额（元）": "门店核销金额", "到店新客数": "门店核销新客数", "到店老客数": "门店核销老客数",
            "视频条数": "门店关联视频数", "同城网吧人气榜排名": "上榜榜单及排名", "门店经营分": "门店经营分",
            "门店评分": "门店评分", "差评率": "经营风险差评率",
            "未成交人数/百分比": "api_visit_lost_count",
            "流失到其他门店成交人数": "api_lost_to_other_store_deal_count",
            "退款金额（元）/百分比": "门店意向退款金额",
        }
        red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')

        # 遍历行来填充数据
        for row in ws.iter_rows(min_row=3): # 从第3行开始遍历
            metric_cell = row[1] # 指标在B列
            metric_name = metric_cell.value
            
            if not metric_name or pd.isna(metric_name):
                continue

            metric_name = str(metric_name).strip()
            
            # 清空旧数据 (跳过F列)
            row[2].value, row[3].value, row[4].value = None, None, None

            if metric_name in field_mapping:
                field_name = field_mapping[metric_name]
                val_current = get_field_value(data_current, field_name)
                val_previous = get_field_value(data_previous, field_name)

                if metric_name in ["商责退单率", "评价回复率", "差评率"]:
                    row[2].value = f"{val_current * 100:.2f}%" if val_current is not None else "数据缺失"
                    row[3].value = f"{val_previous * 100:.2f}%" if val_previous is not None else "数据缺失"
                else:
                    row[2].value = val_current if val_current is not None else "数据缺失"
                    row[3].value = val_previous if val_previous is not None else "数据缺失"
            else:
                log_print(f"--- INFO: 指标 '{metric_name}' 不在映射表中，跳过自动填充。")
                continue

            change_val = calculate_percentage_change(val_current, val_previous)
            row[4].value = change_val if change_val is not None else "数据缺失"
            
            if isinstance(change_val, str) and '↓' in change_val:
                row[4].fill = red_fill
            else:
                row[4].fill = PatternFill(fill_type=None)
        
        # --- 单独处理AI分析写入合并单元格 ---
        log_print("--- DEBUG: [AI分析] 开始生成并写入问题与建议...")
        problem_analysis = generate_ai_analysis(all_data, "problem_analysis")
        improvement_suggestions = generate_ai_analysis(all_data, "improvement_suggestions")
        ai_content = f"【问题分析】\n{problem_analysis}\n\n【改进建议】\n{improvement_suggestions}"
        
        # 写入F3，并设置格式
        analysis_cell = ws['F3']
        analysis_cell.value = ai_content
        analysis_cell.alignment = Alignment(wrap_text=True, vertical='top')
        log_print("--- DEBUG: [AI分析] 已将内容写入F3单元格。")

        # --- 2. 处理“套餐分析”工作表 ---
        write_product_analysis_sheet(wb, all_data, store_name, sort_by, analysis_type)
        
        # --- 3. 保存文件 ---
        log_print(f"--- DEBUG: [报告生成] 准备保存最终的Excel文件: {output_filename}")
        wb.save(output_filename)
        log_print(f"--- SUCCESS: [报告生成] 成功生成Excel报告: {output_filename}")
        return output_filename, f"'{store_name}' 报告生成成功"

    except FileNotFoundError:
        log_print(f"--- FATAL ERROR: [报告生成] 模板文件未找到: '{template_path}'")
        return None, f"错误: 模板文件 '{template_path}' 未找到"
    except Exception as e:
        log_print(f"\n--- EXCEPTION: [报告生成] 失败: {e}"); traceback.print_exc()
        return None, f"错误: 生成报告时发生异常: {e}"

def main_process(store_name, sort_by="销售额", analysis_type="月度分析", data_source="飞书",
                current_start_date=None, current_end_date=None,
                previous_start_date=None, previous_end_date=None,
                current_file=None, previous_file=None):
    """主流程函数，支持多种数据源获取方式。"""
    log_print("\n" + "="*80 + f"\n--- 开始处理门店: {store_name} ({analysis_type}, 数据源: {data_source}) ---\n" + "="*80)
    
    current_store_data = {}
    previous_store_data = {}
    store_id = None
    
    if data_source == "飞书":
        # 原有的飞书数据获取逻辑
        lark_client = lark.Client.builder().app_id(LARK_APP_ID).app_secret(LARK_APP_SECRET).log_level(lark.LogLevel.ERROR).build()
        
        log_print("\n[步骤 1/3] 在线获取门店ID...")
        store_info = get_store_info_from_feishu(store_name, lark_client)
        if not store_info:
            log_print(f"--- 错误: 在飞书表格中找不到名为 '{store_name}' 的门店，请检查名称是否完全匹配！")
            return None, "错误"
            
        store_id_obj = store_info.get("门店ID")
        if isinstance(store_id_obj, list) and store_id_obj and isinstance(store_id_obj[0], dict) and 'text' in store_id_obj[0]:
            store_id = store_id_obj[0]['text']
        else:
            store_id = store_id_obj
        
        if not store_id:
            log_print(f"--- 错误: 无法从飞书记录中解析出有效的门店ID。获取到的值为: {store_id_obj}")
            return None, "错误"
        log_print(f"--- 成功: 查找到门店ID -> {store_id}")

        # 根据分析类型设置日期范围
        current_year = datetime.now().year
        
        if analysis_type == "周分析":
            today = datetime.now()
            weekday = today.weekday()
            current_monday = today - timedelta(days=weekday)
            monday_current = current_monday - timedelta(days=7)
            sunday_current = current_monday - timedelta(days=1)
            monday_previous = current_monday - timedelta(days=14)
            sunday_previous = current_monday - timedelta(days=8)
            
            start_date_current = monday_current
            end_date_current = sunday_current
            start_date_previous = monday_previous
            end_date_previous = sunday_previous

            start_date_current_str = start_date_current.strftime("%Y-%m-%d")
            end_date_current_str = end_date_current.strftime("%Y-%m-%d")
            start_date_previous_str = start_date_previous.strftime("%Y-%m-%d")
            end_date_previous_str = end_date_previous.strftime("%Y-%m-%d")
        else:  # 月度分析
            start_date_current_str = f"{current_year}-09-01"
            end_date_current_str = f"{current_year}-09-30"
            start_date_previous_str = f"{current_year}-08-01"
            end_date_previous_str = f"{current_year}-08-31"

        log_print(f"--- LOG: API查询日期范围 (本期): {start_date_current_str} to {end_date_current_str}")
        log_print(f"--- LOG: API查询日期范围 (上期): {start_date_previous_str} to {end_date_previous_str}")

        # 根据分析类型选择表格ID
        if analysis_type == "周分析":
            current_table_id = LARK_TABLE_ID_WEEK_CURRENT
            previous_table_id = LARK_TABLE_ID_WEEK_PREVIOUS
            product_current_table_id = LARK_TABLE_ID_PRODUCT_WEEK_CURRENT
            product_previous_table_id = LARK_TABLE_ID_PRODUCT_WEEK_PREVIOUS
            current_period_name = "本周数据"
            previous_period_name = "上周数据"
        else:  # 月度分析
            current_table_id = LARK_TABLE_ID_MONTH_CURRENT
            previous_table_id = LARK_TABLE_ID_MONTH_PREVIOUS
            product_current_table_id = LARK_TABLE_ID_PRODUCT_MONTH_CURRENT
            product_previous_table_id = LARK_TABLE_ID_PRODUCT_MONTH_PREVIOUS
            current_period_name = "10月数据"
            previous_period_name = "9月数据"

        log_print("\n[步骤 2/3] 获取门店详细经营数据...")
        results = {}
        filter_query = f"CurrentValue.[门店名称]=\"{store_name}\""
        threads = {
            "current_period_data": threading.Thread(target=lambda: results.update({"current_period_data": query_data_from_table(lark_client, current_table_id, filter_query, current_period_name)})),
            "previous_period_data": threading.Thread(target=lambda: results.update({"previous_period_data": query_data_from_table(lark_client, previous_table_id, filter_query, previous_period_name)})),
            "product_sales_data_current": threading.Thread(target=lambda: results.update({"product_sales_data_current": query_data_from_table(lark_client, product_current_table_id, f"CurrentValue.[核销门店]=\"{store_name}\"", f"{current_period_name}商品")})),
            "product_sales_data_previous": threading.Thread(target=lambda: results.update({"product_sales_data_previous": query_data_from_table(lark_client, product_previous_table_id, f"CurrentValue.[核销门店]=\"{store_name}\"", f"{previous_period_name}商品")})),
        }
        for t in threads.values(): t.start()
        for t in threads.values(): t.join()
        log_print("--- LOG: 所有飞书数据采集任务已结束。")

        current_store_data = results.get("current_period_data")[0] if results.get("current_period_data") else {}
        previous_store_data = results.get("previous_period_data")[0] if results.get("previous_period_data") else {}

        # 从新API获取流失数据
        log_print("\n[步骤 2.5/3] 从外部API获取流失数据...")
        loss_data_current = get_loss_data_from_api(store_id, start_date_current_str, end_date_current_str)
        if loss_data_current:
            current_store_data.update(loss_data_current)
            log_print(f"--- LOG: 已将本期流失数据合并到主数据中。")

        loss_data_previous = get_loss_data_from_api(store_id, start_date_previous_str, end_date_previous_str)
        if loss_data_previous:
            previous_store_data.update(loss_data_previous)
            log_print(f"--- LOG: 已将上期流失数据合并到主数据中。")

        product_sales_data_current = results.get("product_sales_data_current", [])
        product_sales_data_previous = results.get("product_sales_data_previous", [])
        
    elif data_source == "Excel文件":
        # 新的Excel文件数据获取逻辑
        log_print("\n[步骤 1/3] 解析Excel文件数据...")
        
        if current_file:
            current_store_data = parse_excel_data(current_file, store_name)
            if not current_store_data:
                log_print(f"--- 错误: 无法从本期Excel文件中获取门店 '{store_name}' 的数据")
                return None, "错误"
        
        if previous_file:
            previous_store_data = parse_excel_data(previous_file, store_name)
            if not previous_store_data:
                log_print(f"--- 错误: 无法从上期Excel文件中获取门店 '{store_name}' 的数据")
                return None, "错误"
        
        # 对于Excel文件方式，暂时没有商品销售数据和流失数据
        product_sales_data_current = []
        product_sales_data_previous = []
        
    elif data_source == "API下载":
        # API下载数据逻辑
        log_print("\n[步骤 1/3] 从API下载数据...")
        
        # 下载本期数据
        if current_start_date and current_end_date:
            current_file = download_weekly_data(current_start_date, current_end_date)
            if current_file:
                current_store_data = parse_excel_data(current_file, store_name)
                if not current_store_data:
                    log_print(f"--- 警告: 无法从下载的本期数据中获取门店 '{store_name}' 的数据")
        # 下载上期数据
        if previous_start_date and previous_end_date:
            previous_file = download_weekly_data(previous_start_date, previous_end_date)
            if previous_file:
                previous_store_data = parse_excel_data(previous_file, store_name)
                if not previous_store_data:
                    log_print(f"--- 警告: 无法从下载的上期数据中获取门店 '{store_name}' 的数据")
        
        product_sales_data_current = []
        product_sales_data_previous = []

    if not current_store_data and not previous_store_data:
        log_print(f"错误: 未能从任何数据源获取到数据。")
        return None, "错误"

    all_data = {
        "current_period_data": current_store_data,
        "previous_period_data": previous_store_data,
        "products_aggregated": generate_product_insights_with_ai(product_sales_data_current),
        "product_sales_data_current": product_sales_data_current,
        "product_sales_data_previous": product_sales_data_previous,
    }
    
    log_print("\n[步骤 3/3] 生成最终报告...")
    report_path, message = generate_report(store_name, store_id, all_data, sort_by, analysis_type)
    if report_path:
        log_print(f"\n[SUCCESS] 门店 '{store_name}' 任务完成！分析报告已保存为 '{report_path}'！")
    else:
        log_print(f"[ERROR] 错误: 无法为 '{store_name}' 生成报告。")
    return report_path, message

def calculate_week_dates(selected_date):
    """
    根据选择的日期计算自然周范围
    """
    if not selected_date:
        return "", "", "", ""
    
    try:
        # 将字符串转换为日期对象
        if isinstance(selected_date, str):
            selected_date = datetime.strptime(selected_date, "%Y-%m-%d")
        
        # 获取选择日期是星期几 (0=周一, 6=周日)
        weekday = selected_date.weekday()
        
        # 计算选择日期所在周的周一和周日
        monday = selected_date - timedelta(days=weekday)
        sunday = monday + timedelta(days=6)
        
        # 计算上一周的周一和周日
        previous_monday = monday - timedelta(days=7)
        previous_sunday = sunday - timedelta(days=7)
        
        # 格式化日期字符串
        current_start = monday.strftime("%Y-%m-%d")
        current_end = sunday.strftime("%Y-%m-%d")
        previous_start = previous_monday.strftime("%Y-%m-%d")
        previous_end = previous_sunday.strftime("%Y-%m-%d")
        
        return current_start, current_end, previous_start, previous_end
        
    except Exception as e:
        log_print(f"--- ERROR: 日期计算错误: {e}")
        return "", "", "", ""

# ==============================================================================
# --- 5. Gradio 界面 ---
# ==============================================================================

def gradio_interface_function(store_name, analysis_type, sort_by, data_source,
                           current_start_date, current_end_date,
                           previous_start_date, previous_end_date,
                           current_file, previous_file):
    """Gradio界面的主函数，调用核心逻辑并返回结果。"""
    GLOBAL_LOGS.clear()  # 为每次运行重置日志
    
    if not store_name:
        log_print("--- 错误: 请输入门店名称！")
        return None, "\n".join(GLOBAL_LOGS)

    # 根据数据源进行参数验证
    if data_source == "Excel文件":
        if not current_file and not previous_file:
            log_print("--- 错误: 使用Excel文件数据源时，至少需要上传一个文件！")
            return None, "\n".join(GLOBAL_LOGS)
    elif data_source == "API下载":
        if not current_start_date or not current_end_date:
            log_print("--- 错误: 使用API下载数据源时，必须指定本期日期范围！")
            return None, "\n".join(GLOBAL_LOGS)

    # 调用核心处理流程
    report_path, message = main_process(
        store_name, sort_by, analysis_type, data_source,
        current_start_date, current_end_date,
        previous_start_date, previous_end_date,
        current_file, previous_file
    )
    
    # 将最终状态消息也添加到日志中
    if message:
        log_print(f"\n--- 最终状态 ---\n{message}")

    log_output = "\n".join(GLOBAL_LOGS)
    
    # 如果报告路径无效（例如生成失败），返回None
    if not report_path or not os.path.exists(report_path):
        return None, log_output

    return report_path, log_output

if __name__ == "__main__":
    with gr.Blocks() as iface:
        gr.Markdown("# 竞潮玩门店数据分析报告自动生成")
        gr.Markdown("支持多种数据源：飞书API、Excel文件上传、API自动下载")
        
        with gr.Row():
            with gr.Column():
                store_name = gr.Textbox(label="门店名称", placeholder="请输入完整的门店名称...")
                data_source = gr.Radio(
                    choices=["飞书", "Excel文件", "API下载"],
                    label="数据源选择",
                    value="飞书"
                )
                analysis_type = gr.Radio(
                    choices=["月度分析", "周分析"],
                    label="分析维度",
                    value="月度分析"
                )
                sort_by = gr.Radio(
                    choices=["销售额", "核销次数"],
                    label="套餐分析排序方式",
                    value="销售额"
                )
            
            with gr.Column():
                # Excel文件上传区域
                with gr.Group(visible=False) as excel_group:
                    gr.Markdown("### Excel文件上传")
                    current_file = gr.File(label="本期数据文件", file_types=[".xlsx", ".xls"])
                    previous_file = gr.File(label="上期数据文件", file_types=[".xlsx", ".xls"])
                
                # API下载日期选择区域
                with gr.Group(visible=False) as api_group:
                    gr.Markdown("### API下载日期设置")
                    with gr.Row():
                        current_date_picker = gr.Textbox(label="选择本期任意日期（自动计算自然周）", placeholder="YYYY-MM-DD")
                    with gr.Row():
                        current_start_date = gr.Textbox(label="本期开始日期", interactive=False)
                        current_end_date = gr.Textbox(label="本期结束日期", interactive=False)
                    with gr.Row():
                        previous_start_date = gr.Textbox(label="上期开始日期", interactive=False)
                        previous_end_date = gr.Textbox(label="上期结束日期", interactive=False)
        
        # 提交按钮和输出
        submit_btn = gr.Button("生成分析报告", variant="primary")
        
        with gr.Row():
            report_file = gr.File(label="下载分析报告")
            log_output = gr.Textbox(label="运行日志", lines=20, interactive=False)
        
        # 根据数据源选择显示/隐藏相应控件
        def update_visibility(data_source_choice):
            if data_source_choice == "Excel文件":
                return gr.update(visible=True), gr.update(visible=False)
            elif data_source_choice == "API下载":
                return gr.update(visible=False), gr.update(visible=True)
            else:  # 飞书
                return gr.update(visible=False), gr.update(visible=False)
        
        data_source.change(
            fn=update_visibility,
            inputs=[data_source],
            outputs=[excel_group, api_group]
        )
        
        # 日期自动计算事件
        def on_date_change(selected_date):
            if selected_date:
                current_start, current_end, previous_start, previous_end = calculate_week_dates(selected_date)
                return current_start, current_end, previous_start, previous_end
            return "", "", "", ""
        
        current_date_picker.change(
            fn=on_date_change,
            inputs=[current_date_picker],
            outputs=[current_start_date, current_end_date, previous_start_date, previous_end_date]
        )
        
        # 提交事件
        submit_btn.click(
            fn=gradio_interface_function,
            inputs=[
                store_name, analysis_type, sort_by, data_source,
                current_start_date, current_end_date,
                previous_start_date, previous_end_date,
                current_file, previous_file
            ],
            outputs=[report_file, log_output]
        )
    
    iface.launch()
